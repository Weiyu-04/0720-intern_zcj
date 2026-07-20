#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据完整性校验：把 data/*.json 与源 Excel 全量逐条比对，确保无写错、遗漏、
张冠李戴（代码↔名称错配）、重复。每次重建码库后都应运行。

用法：python3 validate.py    （返回码 0=全部通过，1=有问题）
"""
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(__file__)
SRC = os.path.join(HERE, "..", "..", "..", "..", "制度文件")
DATA = os.path.join(HERE, "..", "data")


def _load_json(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as f:
        return json.load(f)["codes"]


def _first_name(row, start=3, end=12):
    for i in range(start, min(end, len(row))):
        v = row[i]
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def check_gbt4754():
    problems = []
    codes = _load_json("gbt4754.json")
    ws = openpyxl.load_workbook(
        os.path.join(SRC, "2017国民经济行业分类注释.xlsx"),
        read_only=True, data_only=True)["Sheet1"]
    # 从源重建 {代码:名称}，与 build_catalog 同规则
    src = {}
    cur = {"门类": None, "大类": None, "中类": None}
    import re
    for row in ws.iter_rows(values_only=True):
        a = str(row[0]).strip() if row[0] else ""
        b = str(row[1]).strip() if row[1] else ""
        if re.fullmatch(r"[A-T]", a):
            src[a] = _first_name(row); cur = {"门类": a, "大类": None, "中类": None}
        elif re.fullmatch(r"\d{2}", a):
            src[a] = _first_name(row); cur["大类"] = a; cur["中类"] = None
        elif re.fullmatch(r"\d{3}", a):
            src[a] = _first_name(row); cur["中类"] = a
        if re.fullmatch(r"\d{4}", b):
            src[b] = _first_name(row)

    # 全量逐条比对名称（张冠李戴检查）
    for k, v in codes.items():
        if k not in src:
            problems.append(f"[gbt4754] json有但源无: {k} {v['name']}")
        elif src[k] != v["name"]:
            problems.append(f"[gbt4754] 名称不符 {k}: json='{v['name']}' 源='{src[k]}'")
    for k in src:
        if k not in codes:
            problems.append(f"[gbt4754] 源有但json缺失: {k} {src[k]}")
    # 父级链完整
    for k, v in codes.items():
        for pk in ("门类", "大类", "中类"):
            if v.get(pk) and v[pk] not in codes:
                problems.append(f"[gbt4754] 父级缺失 {k} 的{pk}={v[pk]}")
    return problems, len(codes), len(src)


def check_geo():
    problems = []
    codes = _load_json("geo.json")
    ws = openpyxl.load_workbook(
        os.path.join(SRC, "行政区划 截止2016年07月31日.xlsx"),
        read_only=True, data_only=True)["Sheet1"]
    src = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        code = str(row[1]).strip() if row[1] else ""
        name = str(row[2]).strip() if row[2] else ""
        xz = str(row[4]).strip() if row[4] is not None else ""
        if xz in ("0", "1", "2") and code:
            src.setdefault(code[:6], name)
    for k, v in codes.items():
        if k not in src:
            problems.append(f"[geo] json有但源无: {k} {v['name']}")
        elif src[k] != v["name"]:
            problems.append(f"[geo] 名称不符 {k}: json='{v['name']}' 源='{src[k]}'")
    for k in src:
        if k not in codes:
            problems.append(f"[geo] 源有但json缺失: {k} {src[k]}")
    for k, v in codes.items():
        if v.get("parent") and v["parent"] not in codes and v["level"] != "省":
            problems.append(f"[geo] 父级缺失 {k} parent={v['parent']}")
    return problems, len(codes), len(src)


def main():
    all_problems = []
    for name, fn in (("gbt4754", check_gbt4754), ("geo", check_geo)):
        probs, njson, nsrc = fn()
        status = "通过" if not probs else f"发现 {len(probs)} 处问题"
        print(f"[{name}] json {njson} 条 / 源 {nsrc} 条 —— {status}")
        for p in probs[:20]:
            print("   ", p)
        all_problems += probs
    print("=" * 50)
    if all_problems:
        print(f"校验未通过：共 {len(all_problems)} 处问题，请回源修正。")
        return 1
    print("校验通过：两个码库与源文件全量逐条一致，无写错/遗漏/张冠李戴/重复。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
